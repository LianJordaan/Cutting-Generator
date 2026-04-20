import shutil
import xlrd
import xlwt
import os
from xlutils.copy import copy as xl_copy
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import warnings
from config_utils import *
from tqdm import tqdm
import re

valid_bord_types = {"plain boards", "grain boards"}
invalid_bord_names = {"own peen", "own grain", "top 600", "top 900"}

warnings.filterwarnings("ignore", message="Workbook contains no default style, apply openpyxl's default")

boardNumbersToRecheck = []

def get_customer_name(file_path):
    return get_cell_value(file_path, 1, 2)

def get_job_name(file_path):
    return get_cell_value(file_path, 2, 2)

def is_safe_filename(name):
    """Return True if filename does NOT contain invalid Windows characters, else False."""
    return not bool(re.search(r'[\\/:"*?<>|]+', name))

def copy_template_to_input_folder(file_path, template_path):
    """Copy template file into the input file's folder and return the new path."""
    input_folder = os.path.dirname(file_path)
    template_filename = os.path.basename(template_path)
    output_path = os.path.join(input_folder, template_filename)

    shutil.copyfile(template_path, output_path)
    return output_path

def process_excel(file_path, template_path):
    print(f"[INFO] Starting processing of file: {file_path}")

    # Extract job name and customer
    print("[INFO] Reading job information from input file...")
    customer = get_customer_name(file_path)
    job_name = get_job_name(file_path)
    print(f"[INFO] Customer: {customer}, Job Name: {job_name}")

    if not is_safe_filename(job_name) or not is_safe_filename(customer):
        raise ValueError(
            f"[ERROR] Invalid characters detected in job name or customer name.\n"
            f"[ERROR] Invalid name(s): Job='{job_name}', Customer='{customer}'\n"
            "Please remove any of the following characters: \\ / : * ? \" < > |"
        )

    output_filename = f"{job_name}.xls"
    output_path = os.path.join(os.path.dirname(file_path), output_filename)

    # Copy template
    print(f"[INFO] Copying template '{template_path}' to '{output_path}'...")
    shutil.copyfile(template_path, output_path)
    print("[INFO] Template copy complete.")

    # Count valid sheets
    num_valid_sheets = count_valid_sheets(file_path)
    print(f"[INFO] Number of valid sheets detected: {num_valid_sheets}")

    # Open template for editing
    print("[INFO] Opening template for modification...")
    rb = xlrd.open_workbook(output_path, formatting_info=True)
    wb = xl_copy(rb)
    print("[INFO] Template loaded and ready for editing.")

    # Prepare print sheet
    wb_print = xlwt.Workbook()
    ws_print = wb_print.add_sheet("Print", True)
    print("[INFO] Print sheet created.")

    current_row_print = 0
    to_write = []

    style_border_all_thin = make_border_style(1, 1, 1, 1)
    style_border_all_thin_bold = make_border_style(1, 1, 1, 1, bold=True)

    # Write initial job info to print sheet
    print("[INFO] Adding job header to print sheet...")
    to_write.append((current_row_print, 0, "Job Name", style_border_all_thin))
    to_write.append((current_row_print, 2, job_name, style_border_all_thin_bold))
    current_row_print += 2

    reading_sheet_index = 0
    processed_sheets = 0

    # Process each valid sheet
    for sheet_index in range(min(get_sheet_count(file_path), rb.nsheets)):
        print(f"[INFO] Processing sheet {sheet_index + 1}/{num_valid_sheets}")
        ws = wb.get_sheet(processed_sheets)
        if not is_sheet_valid(file_path, sheet_index):
            print(f"[WARNING] Sheet {sheet_index + 1} is not valid, skipping...")
            continue

        # Write job info to template
        print(f"[INFO] Writing job name and customer to template sheet {sheet_index + 1}...")
        ws.write(1, 2, job_name, style_border_all_thin)
        ws.write(0, 2, customer, style_border_all_thin)

        # Read bord and edging info
        bord_type = get_cell_value(file_path, 3, 2, sheet_index)
        bord_name = get_cell_value(file_path, 3, 4, sheet_index)
        edging_type = get_cell_value(file_path, 4, 2, sheet_index)
        edging_color = get_cell_value(file_path, 4, 4, sheet_index)
        print(f"[INFO] Bord Type: {bord_type}, Bord Name: {bord_name}")
        print(f"[INFO] Edging Type: {edging_type}, Edging Color: {edging_color}")

        ws.write(2, 2, bord_type, style_border_all_thin)
        ws.write(2, 4, bord_name, style_border_all_thin)
        ws.write(3, 2, edging_type, make_border_style(1, 1, 5, 0))
        ws.write(3, 4, edging_color, make_border_style(1, 0, 5, 1))

        # Fill table data
        last_empty_row = get_last_nonempty_row(file_path, 1, 7, sheet_index)
        print(f"[INFO] Last non-empty row in sheet {sheet_index + 1}: {last_empty_row}")

        total_cells = (last_empty_row - 6) * 7  # Total cells to process (rows × columns)
        progress_bar = tqdm(total=total_cells, desc=f"📄 Processing data for sheet {sheet_index + 1}")

        for loop_col in range(1, 8):
            for loop_row in range(7, last_empty_row + 1):
                value = get_cell_value(file_path, loop_row, loop_col, sheet_index)
                ws.write(loop_row - 1, loop_col, value, style_border_all_thin)
                
                progress_bar.update(1)

        progress_bar.close()
        print(f"[INFO] Table data written to template sheet {sheet_index + 1}")


        # Process edging information
        unique_edging = []
        print("🔧 Processing edging information...")
        for loop_row in tqdm(range(7, last_empty_row + 1), desc="Edging rows"):
            loop_edging_category = str(get_cell_value(file_path, loop_row, 8, sheet_index)).lower()
            loop_edging_name = str(get_cell_value(file_path, loop_row, 9, sheet_index)).lower()

            loop_remark = ""
            if get_cell_value(file_path, loop_row, 10, sheet_index) is not None:
                loop_remark = str(get_cell_value(file_path, loop_row, 10, sheet_index)).lower()

            # Normalization
            loop_edging_category = loop_edging_category.replace("pvc", "")
            loop_edging_category = loop_edging_category.replace("0.4mm", "pvc")
            loop_edging_category = loop_edging_category.replace("3mm", "2x36")

            edging_string = f"{loop_edging_category} {loop_edging_name}".upper()
            edging_string_remark = f"{loop_edging_category} {loop_edging_name} {loop_remark}".upper()

            if "NO EDGING" in edging_string:
                edging_string = "NO EDGING"
            
            if "NO EDGING" in edging_string_remark:
                edging_string_remark = "NO EDGING"

            ws.write(loop_row - 1, 8, loop_edging_category.upper(), style_border_all_thin)
            ws.write(loop_row - 1, 9, loop_edging_name.upper(), style_border_all_thin)    
            ws.write(loop_row - 1, 10, edging_string_remark, style_border_all_thin)

            if edging_string not in unique_edging:
                unique_edging.append(edging_string)
        print(f"[INFO] Unique edgings for sheet {sheet_index + 1}: {unique_edging}")

        # Add cutlist headers to print sheet
        to_write.append((current_row_print, 0, f"Cutlist {sheet_index + 1}", style_border_all_thin))
        to_write.append((current_row_print, 1, "Bord", style_border_all_thin))
        to_write.append((current_row_print, 2, bord_name, style_border_all_thin_bold))
        to_write.append((current_row_print, 3, "Hoeveelheid", style_border_all_thin))
        current_row_print += 3

        # Add edging info to print sheet
        to_write.append((current_row_print, 0, "Edging", style_border_all_thin))
        for edging in unique_edging:
            to_write.append((current_row_print, 2, edging, style_border_all_thin_bold))
            current_row_print += 1
        current_row_print += 1
        
        processed_sheets += 1

    to_write.append((current_row_print, 0, "CUTOUTS", style_border_all_thin))

    # Convert list to dict for easy writing
    to_write_dict = {}
    for row, col, text, style in to_write:
        to_write_dict[(row, col)] = (text, style)

    # Write to print sheet with borders
    print("[INFO] Writing print sheet data with borders...")
    for i in range(5):
        for j in range(current_row_print + 1):
            if (j, i) in to_write_dict:
                text, style = to_write_dict[(j, i)]
                ws_print.write(j, i, text, style)
            else:
                ws_print.write(j, i, "", style_border_all_thin)

    # Set column widths
    print("[INFO] Setting column widths for print sheet...")
    set_column_width_px(ws_print, 0, 70)
    set_column_width_px(ws_print, 2, 200)
    set_column_width_px(ws_print, 3, 88)

    # Save template sheet
    print(f"[INFO] Saving modified template as {output_path}...")
    wb.save(output_path)

    # Save print sheet
    output_path_print = os.path.join(os.path.dirname(file_path), "PRINT " + output_filename.replace('.xls', '_print.xls'))
    print(f"[INFO] Saving print sheet as {output_path_print}...")
    wb_print.save(output_path_print)
    print("[INFO] Processing complete.")

    return boardNumbersToRecheck


def set_column_width_px(ws, col_idx, pixels):
    ws.col(col_idx).width = int((pixels) / 7 * 256)

def get_cell_value(file_path, row, col, sheet_index=0):
    """
    Returns the value at (row, col) from the specified sheet of the input file.
    Row and col are zero-based. Sheet index is zero-based.
    """

    if file_path.lower().endswith('.xlsx'):
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.worksheets[sheet_index]
        # openpyxl uses 1-based indexing for cell access
        cell_value = ws.cell(row=row + 1, column=col + 1).value  # pyright: ignore[reportOptionalMemberAccess]
        return cell_value
    else:
        book = xlrd.open_workbook(file_path)
        sheet = book.sheet_by_index(sheet_index)
        return sheet.cell_value(row, col)

def get_sheet_count(file_path):
    """
    Returns the number of sheets in the input file.
    """
    if file_path.lower().endswith('.xlsx'):
        wb = load_workbook(file_path, read_only=True, data_only=True)
        return len(wb.worksheets)
    else:
        book = xlrd.open_workbook(file_path)
        return book.nsheets

def is_sheet_valid(file_path, sheet_index=0):
    """
    Returns True if the sheet at sheet_index in the input file is valid.
    A valid sheet has 'Plain Boards' or 'Grain Boards' in cell C4 and a non-empty name in E4.
    """
    valid_types = valid_bord_types
    invalid_names = invalid_bord_names

    global boardNumbersToRecheck

    board_category = "Unknown"
    board_name = "Unknown"

    if file_path.lower().endswith('.xlsx'):
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.worksheets[sheet_index]
        value = ws['C4'].value
        board_category = str(value).strip() if value else "Unknown"
        name_value = ws['E4'].value
        board_name = str(name_value).strip() if name_value else "Unknown"
        if isinstance(value, str) and value.strip().lower() in valid_types:
            name_value = ws['E4'].value
            if isinstance(name_value, str) and name_value.strip().lower() not in invalid_names:
                return True
    else:
        book = xlrd.open_workbook(file_path)
        sheet = book.sheet_by_index(sheet_index)
        value = sheet.cell_value(2, 2)
        board_category = str(value).strip() if value else "Unknown"
        name_value = sheet.cell_value(2, 4)
        board_name = str(name_value).strip() if name_value else "Unknown"
        if isinstance(value, str) and value.strip().lower() in valid_types:
            name_value = sheet.cell_value(2, 4)
            if isinstance(name_value, str) and name_value.strip().lower() not in invalid_names:
                return True
    boardNumbersToRecheck.append((sheet_index + 1, board_category, board_name))
    return False

def count_valid_sheets(file_path):
    """
    Returns the number of sheets where cell C3 is 'Plain Boards' or 'Grain Boards'.
    """
    valid_types = valid_bord_types
    invalid_names = invalid_bord_names
    count = 0

    if file_path.lower().endswith('.xlsx'):
        wb = load_workbook(file_path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            value = ws['C4'].value
            if isinstance(value, str) and value.strip().lower() in valid_types:
                name_value = ws['E4'].value
                if isinstance(name_value, str) and name_value.strip().lower() not in invalid_names:
                    count += 1
    else:
        book = xlrd.open_workbook(file_path)
        for sheet in book.sheets():
            value = sheet.cell_value(2, 2)
            if isinstance(value, str) and value.strip().lower() in valid_types:
                name_value = sheet.cell_value(2, 4)
                if isinstance(name_value, str) and name_value.strip().lower() not in invalid_names:
                    count += 1
    return count

def get_last_nonempty_row(file_path, col, start_row, sheet_index=0):
    """
    Returns the row index (zero-based) of the last non-empty cell in the given column,
    starting from start_row, in the specified sheet.
    """
    if file_path.lower().endswith('.xlsx'):
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.worksheets[sheet_index]
        max_row = ws.max_row
        last_row = None
        for r in range(start_row + 1, max_row + 1):  # openpyxl is 1-based
            value = ws.cell(row=r, column=col + 1).value
            if value not in (None, ""):
                last_row = r - 1  # convert to zero-based
        return last_row
    else:
        book = xlrd.open_workbook(file_path)
        sheet = book.sheet_by_index(sheet_index)
        nrows = sheet.nrows
        last_row = None
        for r in range(start_row, nrows):
            value = sheet.cell_value(r, col)
            if value not in ("", None):
                last_row = r
        return last_row

def make_border_style(top, right, bottom, left, font_name='Calibri', font_size=12, bold=False):
    """
    Returns an xlwt.XFStyle with the specified border styles, Calibri font, and font size 11.
    Border values: 0=No border, 1=Thin, 2=Medium, 3=Dashed, 4=Dotted, 5=Thick, etc.
    """
    style = xlwt.XFStyle()
    style.borders.top = top
    style.borders.right = right
    style.borders.bottom = bottom
    style.borders.left = left
    font = xlwt.Font()
    font.name = font_name
    font.height = font_size * 20  # xlwt uses twips (1/20 of a point)
    font.bold = bold
    style.font = font
    return style

import fdb

def find_cutouts(quote_nr):
    config = load_config()
    ip = config.get("ip", None)
    port = config.get("port", "3050")
    username = config.get("username", None)
    password = config.get("password", None)
    filepath = config.get("filepath", "C:/ZAWare/DB/CutMan/CUTMAN.FDB")
    charset = config.get("charset", "UTF8")

    if not ip or not username or not password:
        print("[ERROR] Database configuration is incomplete. Please run the setup.")
        return []

    con = fdb.connect(
        dsn=f'{ip}/{port}:{filepath}',
        user=f'{username}',
        password=f'{password}',
        charset=f'{charset}'
    )
    cur = con.cursor()

    print("[INFO] Searcing cutlist ids for quote number:", quote_nr)
    # Get cutlist_ids
    cur.execute("""
        SELECT CUTLIST_ID
        FROM CUTLIST
        WHERE QUOTE_NR = ?
    """, (quote_nr,))
    cutlist_ids = [row[0] for row in cur.fetchall()]
    if not cutlist_ids:
        cur.close()
        con.close()
        return []

    print("[INFO] Found cutlist IDs:", cutlist_ids)
    print("[INFO] Fetching cutout details...")

    results = []
    for cutlist_id in cutlist_ids:
        print("[INFO] Processing cutlist ID:", cutlist_id)
        cur.execute("""
            SELECT ITEM_ID, LENGTE, WYDTE, QTY, BOARD_ID
            FROM CUT_LIST_DETAIL
            WHERE QUOTE_NR = ? AND CUTLIST_ID = ?
        """, (quote_nr, cutlist_id))
        for item_id, lengte, wydte, qty, board_id in cur.fetchall():
            cur.execute("""
                SELECT CUTOUT1, CUTOUT2
                FROM CUTOUTS
                WHERE QUOTE_NR = ? AND CUTLIST_ID = ? AND ITEM_ID = ?
            """, (quote_nr, cutlist_id, item_id))
            cutout_row = cur.fetchone()
            if cutout_row:
                print("[INFO] Found cutouts for item ID:", item_id)
                print("[INFO] Data for piece with cutout: Length:", lengte, "Width:", wydte, "Qty:", qty, "Cutout Data:", cutout_row)
                cutout1, cutout2 = cutout_row
                board_name = "Unknown"
                cur.execute("""
                    SELECT BOARD_NAME
                    FROM BOARD_TYPES
                    WHERE BOARD_ID = ?
                """, (board_id,))
                board_row = cur.fetchone()
                print("DEBUG: Board ID:", board_id, "Board Row:", board_row, "Type:", type(board_row), cutout_row)
                if board_row:
                    board_name = board_row[0]
                results.append((item_id, cutlist_id, lengte, wydte, qty, cutout1, cutout2, board_name))

    cur.close()
    con.close()
    return results


def find_crosscuts(quote_nr):
    config = load_config()
    ip = config.get("ip", None)
    port = config.get("port", "3050")
    username = config.get("username", None)
    password = config.get("password", None)
    filepath = config.get("filepath", "C:/ZAWare/DB/CutMan/CUTMAN.FDB")
    charset = config.get("charset", "UTF8")

    if not ip or not username or not password:
        print("[ERROR] Database configuration is incomplete. Please run the setup.")
        return []

    con = fdb.connect(
        dsn=f'{ip}/{port}:{filepath}',
        user=f'{username}',
        password=f'{password}',
        charset=f'{charset}'
    )
    cur = con.cursor()

    print("[INFO] Searcing cutlist ids for quote number:", quote_nr)
    # Get cutlist_ids
    cur.execute("""
        SELECT CUTLIST_ID
        FROM CUTLIST
        WHERE QUOTE_NR = ?
    """, (quote_nr,))
    cutlist_ids = [row[0] for row in cur.fetchall()]
    if not cutlist_ids:
        cur.close()
        con.close()
        return []

    print("[INFO] Found cutlist IDs:", cutlist_ids)
    print("[INFO] Fetching crosscut details...")

    results = []
    for cutlist_id in cutlist_ids:
        print("[INFO] Processing cutlist ID:", cutlist_id)
        cur.execute("""
            SELECT ITEM_ID, LENGTE, WYDTE, QTY, BOARD_ID
            FROM CUT_LIST_DETAIL
            WHERE QUOTE_NR = ? AND CUTLIST_ID = ?
        """, (quote_nr, cutlist_id))
        for item_id, lengte, wydte, qty, board_id in cur.fetchall():
            cur.execute("""
                SELECT LENGTE
                FROM CROSSCUTS
                WHERE QUOTE_NR = ? AND CUTLIST_ID = ? AND ITEM_ID = ?
            """, (quote_nr, cutlist_id, item_id))
            crosscut_rows = cur.fetchall()
            if crosscut_rows:
                list_of_lengths = []
                for crosscut in crosscut_rows:
                    list_of_lengths.append(crosscut[0])
                print("[INFO] Data for piece with crosscuts: Length:", lengte, "Width:", wydte, "Qty:", qty, "Crosscut Data:", list_of_lengths)

                board_name = "Unknown"
                cur.execute("""
                    SELECT BOARD_NAME
                    FROM BOARD_TYPES
                    WHERE BOARD_ID = ?
                """, (board_id,))
                board_row = cur.fetchone()
                if board_row:
                    board_name = board_row[0]

                t = (lengte, wydte, qty, board_name)
                extra = list_of_lengths

                # unpack the list so its elements are added, not the list itself
                crosscut_data = (*t, *extra)

                results.append(crosscut_data)

    cur.close()
    con.close()
    return results

def containsSpecialCrosscuts(quote_nr):
    config = load_config()
    ip = config.get("ip", None)
    port = config.get("port", "3050")
    username = config.get("username", None)
    password = config.get("password", None)
    filepath = config.get("filepath", "C:/ZAWare/DB/CutMan/CUTMAN.FDB")
    charset = config.get("charset", "UTF8")

    if not ip or not username or not password:
        print("[ERROR] Database configuration is incomplete. Please run the setup.")
        return []

    con = fdb.connect(
        dsn=f'{ip}/{port}:{filepath}',
        user=f'{username}',
        password=f'{password}',
        charset=f'{charset}'
    )
    cur = con.cursor()

    print("[INFO] Searcing for cutlist ids for quote number:", quote_nr)
    # Get cutlist_ids
    cur.execute("""
        SELECT CUTLIST_ID
        FROM CUTLIST
        WHERE QUOTE_NR = ?
    """, (quote_nr,))
    cutlist_ids = [row[0] for row in cur.fetchall()]
    if not cutlist_ids:
        cur.close()
        con.close()
        return []

    print("[INFO] Found cutlist IDs:", cutlist_ids)
    print("[INFO] Fetching for special crosscuts...")

    results = []
    for cutlist_id in cutlist_ids:
        print("[INFO] Processing cutlist ID:", cutlist_id)
        cur.execute("""
            SELECT ITEM_ID, LENGTE, WYDTE, QTY, BOARD_ID
            FROM CUT_LIST_DETAIL
            WHERE QUOTE_NR = ? AND CUTLIST_ID = ?
        """, (quote_nr, cutlist_id))
        for item_id, lengte, wydte, qty, board_id in cur.fetchall():
            cur.execute("""
                SELECT SPLIT, HOLES_L, HOLES_S, LSWING
                FROM CROSSCUTS
                WHERE QUOTE_NR = ? AND CUTLIST_ID = ? AND ITEM_ID = ?
            """, (quote_nr, cutlist_id, item_id))
            crosscut_rows = cur.fetchall()
            if crosscut_rows:
                print("[INFO] Found special crosscuts.. Processing...")
                return True

    cur.close()
    con.close()
    return False

def normalize_board_types(board_color):
    categoties = ["Plain Boards", "Grain Boards", "Plywood", "Plain Hardboard"]
    # give the user whatever the color is, and then ask the user to enter a number, basically pick what category that one belongs to.
    print(f"Board color: {board_color}")
    print("Please enter the number corresponding to the category this board belongs to:")
    for i, category in enumerate(categoties):
        print(f"{i}. {category}")
    while True:
        choice = input("Enter the number of the category: ")
        if choice.isdigit() and 0 <= int(choice) < len(categoties):
            return categoties[int(choice)]
        else:
            print("Invalid choice. Please enter a valid number.")

def normalize_edging_types(edging):
    # give the user whatever the edging is, and then ask the user to enter a number, basically pick what category that one belongs to.
    categories = ["0.4mm PVC", "1mm PVC", "2mm PVC", "3mm PVC"]
    if edging == "":
        return categories[0]
    print(f"Edging: {edging}")
    print("Please enter the number corresponding to the category this edging belongs to:")
    for i, category in enumerate(categories):
        print(f"{i}. {category}")
    while True:
        choice = input("Enter the number of the category: ")
        if choice.isdigit() and 0 <= int(choice) < len(categories):
            return categories[int(choice)]
        else:
            print("Invalid choice. Please enter a valid number.")

def normalize_extra_data(extra_data):
    text = "" if extra_data is None else str(extra_data).replace("\xa0", " ").strip()

    if text == "":
        return 0, ""

    if "gate" not in text.lower():
        default_text = text.upper()
        print(f"Default extra text: {default_text}")
        user_text = input("Enter replacement extra text (press Enter to keep default): ")
        replacement_text = user_text.upper() if user_text else default_text
        return 0, replacement_text

    # Example input: "Boor 2 x gate: 100mm van hoeke af"
    # Extract gate count (the number before "gate").
    gate_count_match = re.search(r"(\d+)\s*x?\s*gate", text, re.IGNORECASE)
    gate_count = int(gate_count_match.group(1)) if gate_count_match else 0

    # Keep only the text after ":" as the default extra description.
    text_after_colon = text.split(":", 1)[1].strip() if ":" in text else ""
    default_text = text_after_colon.upper()

    print(f"Detected gate count: {gate_count}")
    print(f"Default extra text: {default_text}")
    user_text = input("Enter replacement extra text (press Enter to keep default): ")

    replacement_text = user_text.upper() if user_text else default_text
    return gate_count, replacement_text


def process_erik_cutlist(file_path, template_path):
    def to_int_value(value):
        """Convert numeric-looking values to rounded int; return None when not numeric."""
        if isinstance(value, (int, float)):
            return int(round(value))
        if isinstance(value, str):
            text = value.strip()
            if text == "":
                return None
            try:
                return int(round(float(text.replace(",", "."))))
            except ValueError:
                return None
        return None

    def set_cell_value_safe_for_merge(ws, row, column, value):
        """Write to a cell, redirecting to merged-range anchor when needed."""
        cell = ws.cell(row=row, column=column)
        if not isinstance(cell, MergedCell):
            cell.value = value
            return

        for merged_range in ws.merged_cells.ranges:
            if (
                merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= column <= merged_range.max_col
            ):
                ws.cell(row=merged_range.min_row, column=merged_range.min_col, value=value)
                return

        # Fallback in case no matching range is found
        ws.cell(row=row, column=column, value=value)

    newTemplatePath = copy_template_to_input_folder(file_path, template_path)
    # rename the copied template to "CUTTING - {job_name}.xlsx"
    job_name = os.path.splitext(os.path.basename(file_path))[0]
    new_template_name = f"CUTTING - {job_name}.xlsx"
    new_template_path = os.path.join(os.path.dirname(newTemplatePath), new_template_name)
    os.rename(newTemplatePath, new_template_path)
    # open the input file in read only mode. in the B column, go down untill you find a cell that contains "die einde" ignore case also contain, not exact mach. then save the index. i need to know how far down i will have to search for data.
    book = xlrd.open_workbook(file_path)
    sheet = book.sheet_by_index(0)
    end_row_index = None
    for row_idx in range(sheet.nrows):
        cell_value = sheet.cell_value(row_idx, 1)  # Column B is index 1
        if isinstance(cell_value, str) and "die einde" in cell_value.lower():
            end_row_index = row_idx
            break
    if end_row_index is None:
        print("[ERROR] Could not find 'die einde' in column B. Cannot process Erik cutlist.")
        return
    # print(f"[INFO] Found 'die einde' at row index: {end_row_index}")

    # loop in the A column from the first row all the way down to the end row - 1.
    current_board_color = None
    current_board_category = None
    current_edging_color = None
    current_edging_category = None

    known_board_category_mappings = {}
    known_edging_category_mappings = {}

    known_extra_data_mappings = {}

    all_boards_data = {}

    
    for row_idx in range(end_row_index):
        cell_value = sheet.cell_value(row_idx, 0)
        if "board type" in cell_value.lower():
            current_board_color = sheet.cell_value(row_idx, 1)
            if current_board_color in known_board_category_mappings:
                current_board_category = known_board_category_mappings[current_board_color]
                # print(f"[INFO] Board category for color '{current_board_color}' found in cache: {current_board_category}")
            else:
                current_board_category = normalize_board_types(current_board_color)
                known_board_category_mappings[current_board_color] = current_board_category
            # print(f"[INFO] Board category changed: {current_board_category} (Color: {current_board_color})")
    
    
    for row_idx in range(end_row_index):
        cell_value = sheet.cell_value(row_idx, 0)
        if "edging" in cell_value.lower():
            current_edging_color = sheet.cell_value(row_idx, 1)
            if current_edging_color in known_edging_category_mappings:
                current_edging_category = known_edging_category_mappings[current_edging_color]
                # print(f"[INFO] Edging category for color '{current_edging_color}' found in cache: {current_edging_category}")
            else:
                current_edging_category = normalize_edging_types(current_edging_color)
                known_edging_category_mappings[current_edging_color] = current_edging_category
            # print(f"[INFO] Edging category changed: {current_edging_category} (Color: {current_edging_color})")

    for row_idx in range(end_row_index):
        cell_value = sheet.cell_value(row_idx, 0)
        if "board type" in cell_value.lower():
            current_board_color = sheet.cell_value(row_idx, 1)
            if current_board_color in known_board_category_mappings:
                current_board_category = known_board_category_mappings[current_board_color]
                # print(f"[INFO] Board category for color '{current_board_color}' found in cache: {current_board_category}")
            else:
                current_board_category = normalize_board_types(current_board_color)
                known_board_category_mappings[current_board_color] = current_board_category
            # print(f"[INFO] Board category changed: {current_board_category} (Color: {current_board_color})")
        if "edging" in cell_value.lower():
            current_edging_color = sheet.cell_value(row_idx, 1)
            if current_edging_color in known_edging_category_mappings:
                current_edging_category = known_edging_category_mappings[current_edging_color]
                # print(f"[INFO] Edging category for color '{current_edging_color}' found in cache: {current_edging_category}")
            else:
                current_edging_category = normalize_edging_types(current_edging_color)
                known_edging_category_mappings[current_edging_color] = current_edging_category
            # print(f"[INFO] Edging category changed: {current_edging_category} (Color: {current_edging_color})")
        length = to_int_value(sheet.cell_value(row_idx, 1))
        width = to_int_value(sheet.cell_value(row_idx, 2))
        quantity = to_int_value(sheet.cell_value(row_idx, 3))
        edge_length = to_int_value(sheet.cell_value(row_idx, 4))
        edge_width = to_int_value(sheet.cell_value(row_idx, 5))
        extra = sheet.cell_value(row_idx, 6)

        # check if length and width are both numbers
        if length is not None and width is not None:
            # print(f"[INFO] Processing piece: Length={length}, Width={width}, Quantity={quantity}, Edge Length={edge_length}, Edge Width={edge_width}, Extra={extra}")
            # instead of writing this data to the template file, we want to store it inside a variable, and it should be categorized by per board_color.

            holes = 0

            if current_board_color not in all_boards_data:
                all_boards_data[current_board_color] = []
            if extra is not None and str(extra).replace("\xa0", " ").strip() != "":
                if extra in known_extra_data_mappings:
                    holes, extra = known_extra_data_mappings[extra]
                else:
                    known_extra_data_mappings[extra] = normalize_extra_data(extra)
                    holes, extra = known_extra_data_mappings[extra]
            else:
                extra = ""

            all_boards_data[current_board_color].append({
                "length": length,
                "width": width,
                "quantity": quantity,
                "edge_length": edge_length,
                "edge_width": edge_width,
                "holes": holes,
                "extra": extra,
                "edge_category": current_edging_category,
                "board_category": current_board_category,
                "edge_color": current_edging_color,
                "board_color": current_board_color
            })
    # print("[DEBUG]" + str(all_boards_data))

    current_sheet_index = 0

    # open the new xlsx template file for writing
    wb = load_workbook(new_template_path)
    active_sheet = None

    # loop thru the indexes of all_boards_data
    for board_color, pieces in all_boards_data.items():
        print(f"[INFO] Writing data for board color: {board_color} with {len(pieces)} pieces. Onto sheet index: {current_sheet_index}")
        if current_sheet_index >= len(wb.worksheets):
            print(f"[WARNING] Not enough sheets in template for board color '{board_color}'. Skipping remaining data.")
            break

        active_sheet = wb.worksheets[current_sheet_index]
        
        set_cell_value_safe_for_merge(active_sheet, 4, 3, pieces[0]['board_category'])
        # the cell below that should be the edging category
        set_cell_value_safe_for_merge(active_sheet, 5, 3, pieces[0]['edge_category'])

        set_cell_value_safe_for_merge(active_sheet, 4, 5, pieces[0]['board_color'])
        set_cell_value_safe_for_merge(active_sheet, 5, 5, pieces[0]['edge_color'])

        row_index = 7

        # for each piece, write the data to the template file, but also include the board category and edging category in the print sheet.
        for piece in pieces:
            # print(f"[INFO] Writing piece: Length={piece['length']}, Width={piece['width']}, Quantity={piece['quantity']}, Edge Length={piece['edge_length']}, Edge Width={piece['edge_width']}, Extra={piece['extra']}, Board Category={piece['board_category']}, Edging Category={piece['edge_category']}")

            # print(piece['edge_color'])
            if piece['edge_color'] == "":
                piece['edge_color'] = "** NO EDGING"

            excel_row = row_index + 1
            active_sheet.cell(row=excel_row, column=2, value=piece['length'])
            active_sheet.cell(row=excel_row, column=3, value=piece['width'])
            active_sheet.cell(row=excel_row, column=4, value=piece['quantity'])
            active_sheet.cell(row=excel_row, column=5, value=piece['edge_length'])
            active_sheet.cell(row=excel_row, column=6, value=piece['edge_width'])

            active_sheet.cell(row=excel_row, column=7, value=piece['holes'])

            active_sheet.cell(row=excel_row, column=9, value=piece['edge_category'])
            active_sheet.cell(row=excel_row, column=10, value=piece['edge_color'])
            active_sheet.cell(row=excel_row, column=11, value=piece['extra'])

            row_index += 1

        current_sheet_index += 1
    
    # make it save..?
    wb.save(new_template_path)